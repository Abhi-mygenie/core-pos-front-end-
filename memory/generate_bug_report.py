import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Bug Report Aug 7–14 2026"

# ── Palette ──────────────────────────────────────────────────────────────────
BRAND_GREEN   = "1F6B35"   # MyGenie dark green
BRAND_ORANGE  = "D9480F"   # MyGenie orange
HEADER_BG     = "2D6A4F"   # Deep green header
HEADER_FG     = "FFFFFF"

STATUS_COLORS = {
    "IMPLEMENTED":              ("E8F5E9", "1B5E20"),
    "IMPLEMENTED — QA PASS":    ("C8E6C9", "1B5E20"),
    "BACKEND-BLOCKED":          ("FFF3E0", "E65100"),
    "GATE 3 COMPLETE — Awaiting Gate 4 GO": ("E3F2FD", "0D47A1"),
    "SUBSUMED (Implemented via CR-139)":    ("F3E5F5", "4A148C"),
}

PRIORITY_COLORS = {
    "P0": "B71C1C",
    "P1": "E53935",
    "P2": "FB8C00",
    "P3": "43A047",
}

# ── Data ─────────────────────────────────────────────────────────────────────
bugs = [
    # (ID, Date, Priority, Title/Description, Status)
    ("BUG-SCAN-DEDUP", "2026-08-08", "P1",
     "show_scan_popup loads wrong value & save silently lost — backend dedup (2026-08-08) moved field from advanced{} to basic{}; FE still read/wrote advanced.show_scan_popup",
     "IMPLEMENTED"),

    ("BUG-303", "2026-08-11", "P2",
     "P&L Report — 'Paid Revenue' KPI always shows ₹0 (field mismatch: s.paid_revenue vs s.total_paid_revenue in PLReportPage.jsx)",
     "IMPLEMENTED — QA PASS"),

    ("BUG-304", "2026-08-11", "P1",
     "Item-Level Discount — discountRatio uses full itemTotal (not discountableTotal), causing wrong GST/VAT for non-discountable items (CollectPaymentPanel.jsx + CartPanel.jsx)",
     "IMPLEMENTED — QA PASS"),

    ("BUG-305", "2026-08-11", "P1",
     "orderTransform.js — discountRatio uses full subtotal in calcOrderTotals + buildBillPrintPayload; wrong GST sent to backend payload and bill print (CRITICAL risk)",
     "IMPLEMENTED — QA PASS"),

    ("BUG-306", "2026-08-11", "P1",
     "Aggregator Setup shows 'Network Error' / blank screen when GET /aggregator-config returns ERR_NETWORK (AggregatorSetupView.jsx isNoConfig guard missing)",
     "IMPLEMENTED — QA PASS"),

    ("BUG-307", "2026-08-11", "P1",
     "Aggregator Setup: tone_timing (notification duration seconds) not mapped in UI — field missing from aggregatorConfigTransform.js fromAPI/toAPI + ConfigTab.jsx",
     "IMPLEMENTED — QA PASS"),

    ("BUG-296", "2026-08-12", "P1",
     "Food Court Report vs Item-Wise Report — Revenue data mismatch (rid=598, Shimla QoH Food Court, June 2026). Two root causes: wrong sort_by field (created_at→collect_bill) + foodStatus=3 items not excluded (foodCourtService.js)",
     "IMPLEMENTED — QA PASS"),

    ("BUG-PL-A", "2026-08-12", "P1",
     "P&L Report — API returns all-zeros because reportService.js sends date_from/date_to but endpoint expects from/to parameter keys",
     "IMPLEMENTED"),

    ("BUG-PL-B", "2026-08-12", "P2",
     "P&L Report — API returns comma-formatted number strings (e.g. '537,876.02'); parseFloat() truncates at comma → KPI values cut off (reportService.js / PLReportPage.jsx)",
     "IMPLEMENTED"),

    ("BUG-308", "2026-08-13", "P1",
     "Sub-Recipe Stock: StockAuditPanel calls addStock() (ingredient endpoint) instead of addSubRecipeStock() for sub-recipe entries — wrong API, wrong payload (constants.js + inventoryTransform.js + inventoryService.js + StockAuditPanel.jsx)",
     "IMPLEMENTED"),

    ("BUG-309", "2026-08-13", "P1",
     "Ingredient Bulk Edit: Min Unit column uses <input type=number> which drops the unit string on save — data loss on every save (IngredientBulkEditor.jsx:442 input→span)",
     "IMPLEMENTED"),

    ("BUG-310", "2026-08-13", "P2",
     "Ingredient Bulk Edit: Conversion Factor field has transparent border/background (numCls false) — looks like static text, users cannot tell it is editable (IngredientBulkEditor.jsx:296)",
     "IMPLEMENTED"),

    ("BUG-311", "2026-08-13", "P1",
     "Ingredient Add / Bulk Edit: No duplicate detection at any layer — missing typeahead L1 (position:fixed dropdown in InventorySetupPanel.jsx:23), pre-save isDuplicate guard L2 (InventorySetupPanel.jsx:146), bulk-editor dupe skip L3 (IngredientBulkEditor.jsx:192)",
     "IMPLEMENTED"),

    ("BUG-312", "2026-08-13", "P1",
     "fromAPI.ingredients() missing isSubRecipe/subrecipeId fields — root cause for all sub-recipe misrouting throughout inventory transforms (inventoryTransform.js fromAPI.ingredients())",
     "SUBSUMED (Implemented via CR-139)"),

    ("BUG-313", "2026-08-13", "P1",
     "Sub-recipe rows appear in Stock Update auto-plan + addPurchase() called for all rows with no sub-recipe routing guard (purchasePlanner.js, AutoShoppingList.jsx, PurchaseEntryPanel.jsx, SmartPurchasePanel.jsx)",
     "SUBSUMED (Implemented via CR-139)"),

    ("BUG-314", "2026-08-13", "P1",
     "Inventory Setup: Categories (0) + Unit dropdown empty — Promise.all atomic failure when get-inventory-master returns 404; categories and units never set (InventorySetupPanel.jsx:42 → Promise.allSettled fix)",
     "IMPLEMENTED"),

    ("BUG-315", "2026-08-13", "P2",
     "Printer Config: Numeric inputs snap-back — cannot clear '1' to retype value; controlled input returns '' early → React reverts (StyleInput in PrintStyleTab.jsx + NumberInput in shared.jsx → local display state fix)",
     "IMPLEMENTED"),

    ("BUG-316", "2026-08-13", "P1",
     "Printer Config: Font Family dropdown empty because available_fonts is null from API — printerAgentConfigTransform.js:253 had no fallback; fix: FALLBACK_FONTS constant with 11 approved fonts",
     "IMPLEMENTED"),

    ("BUG-317", "2026-08-13", "P2",
     "Printer Config: Android size fields (Logo/UPI/FeedbackQR) reject values >8 due to max=8 constraint inherited from CR-133 [1,8] default range — PrintStyleTab.jsx maxScale removed",
     "IMPLEMENTED"),

    ("BUG-318", "2026-08-13", "P1",
     "Printer Config: Aggregator auto-print keys (auto_kot/auto_bill/auto_stage) missing from AutoPrintTab UI + saves to wrong API endpoint; CR-133 OD-B misrouted these — owner reversed OD-B; AutoPrintTab.jsx full rewrite + FALLBACK_AGGREGATOR_STAGES added",
     "IMPLEMENTED"),

    ("BUG-319", "2026-08-13", "P2",
     "Printer Config: Footer text 'Powered by MyGenie' hardcoded in print agent firmware — API returns bill_footer.footer_text correctly but physical print agent ignores it; backend firmware change required",
     "BACKEND-BLOCKED"),

    ("BUG-320", "2026-08-13", "P2",
     "Sub-Recipe Stock: physical_qty incorrectly included in add-sub-recipe-stock payload — physical_qty is an ingredient audit concept not applicable to sub-recipe produced-qty; always mirrored quantity (SubRecipeStockPanel.jsx:94 + inventoryTransform.js:227)",
     "IMPLEMENTED"),

    ("BUG-236", "2026-08-14", "P1",
     "Smart Purchase — Ad-hoc Typeahead Dropdown clipped by overflow-hidden on Section 1 card; z-index too low (z-10→z-50). Fix: removed overflow-hidden from card (L130) + z-50 on both dropdown divs (L42, L47) in AdHocTypeahead (SmartPurchasePanel.jsx)",
     "IMPLEMENTED"),

    ("BUG-321", "2026-08-14", "P1",
     "Sub-Recipe Stock Panel — Wrong Produce/Recount semantics: every save adds qty (ADD mode) but UI shows drift/wastage screen; StockAuditPanel.jsx:71 physicalQty=qty causes spurious wastage. Fix: mode toggle (Produce/Recount) + conditional physical_qty in transform + StockAuditPanel routing (3 files)",
     "GATE 3 COMPLETE — Awaiting Gate 4 GO"),

    ("BUG-322", "2026-08-14", "P1",
     "Recipe Form — SearchableSelect ingredient dropdown clipped by overflow-hidden on table container (RecipeFormPanel.jsx:305); position:absolute dropdown hidden behind table. Fix: position:fixed + getBoundingClientRect on trigger. All recipe types affected (standard/sub/addon ingredient rows). Related: BUG-236, BUG-238",
     "IMPLEMENTED"),
]

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = {
    "A": 18,   # Bug ID
    "B": 14,   # Date
    "C": 10,   # Priority
    "D": 90,   # Description
    "E": 40,   # Status
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# ── Title row ─────────────────────────────────────────────────────────────────
ws.merge_cells("A1:E1")
title_cell = ws["A1"]
title_cell.value = "MyGenie POS — Bug Report  |  7 Aug 2026 → 14 Aug 2026"
title_cell.font = Font(name="Calibri", bold=True, size=14, color=HEADER_FG)
title_cell.fill = PatternFill("solid", fgColor=BRAND_GREEN)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# ── Subtitle ──────────────────────────────────────────────────────────────────
ws.merge_cells("A2:E2")
sub = ws["A2"]
sub.value = f"Source: /app/memory/control/BUG_TRACKER.md  ·  Total bugs: {len(bugs)}"
sub.font = Font(name="Calibri", italic=True, size=10, color="444444")
sub.fill = PatternFill("solid", fgColor="F1F8E9")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# ── Header row ────────────────────────────────────────────────────────────────
headers = ["Bug ID", "Date", "Priority", "Description", "Status"]
header_row = 3
for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=col_idx, value=h)
    cell.font = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[header_row].height = 22

thin = Side(style="thin", color="B0BEC5")
thick = Side(style="medium", color="2D6A4F")

def apply_border(cell, top=thin, bottom=thin, left=thin, right=thin):
    cell.border = Border(top=top, bottom=bottom, left=left, right=right)

for col_idx in range(1, 6):
    apply_border(ws.cell(row=header_row, column=col_idx),
                 top=thick, bottom=thick,
                 left=thick if col_idx == 1 else thin,
                 right=thick if col_idx == 5 else thin)

# ── Data rows ─────────────────────────────────────────────────────────────────
for row_offset, (bug_id, date, priority, description, status) in enumerate(bugs):
    r = header_row + 1 + row_offset
    ws.row_dimensions[r].height = 60

    # Alternate row shading
    row_bg = "F9FBF9" if row_offset % 2 == 0 else "FFFFFF"

    # Resolve status display colors
    status_bg, status_fg = STATUS_COLORS.get(status, ("FFFFFF", "000000"))

    # Priority color
    pri_color = PRIORITY_COLORS.get(priority, "333333")

    values = [bug_id, date, priority, description, status]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=r, column=col_idx, value=val)
        cell.alignment = Alignment(
            horizontal="center" if col_idx != 4 else "left",
            vertical="top",
            wrap_text=True
        )
        apply_border(cell,
                     left=thick if col_idx == 1 else thin,
                     right=thick if col_idx == 5 else thin)

        # Bug ID column
        if col_idx == 1:
            cell.font = Font(name="Calibri", bold=True, size=10, color=BRAND_GREEN)
            cell.fill = PatternFill("solid", fgColor=row_bg)

        # Date column
        elif col_idx == 2:
            cell.font = Font(name="Calibri", size=10, color="546E7A")
            cell.fill = PatternFill("solid", fgColor=row_bg)

        # Priority column
        elif col_idx == 3:
            cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=pri_color)

        # Description column
        elif col_idx == 4:
            cell.font = Font(name="Calibri", size=9, color="212121")
            cell.fill = PatternFill("solid", fgColor=row_bg)

        # Status column
        elif col_idx == 5:
            cell.font = Font(name="Calibri", bold=True, size=9, color=status_fg)
            cell.fill = PatternFill("solid", fgColor=status_bg)

# ── Bottom border on last data row ────────────────────────────────────────────
last_row = header_row + len(bugs)
for col_idx in range(1, 6):
    cell = ws.cell(row=last_row, column=col_idx)
    cell.border = Border(
        top=thin, bottom=thick,
        left=thick if col_idx == 1 else thin,
        right=thick if col_idx == 5 else thin
    )

# ── Legend ────────────────────────────────────────────────────────────────────
legend_row = last_row + 2
ws.merge_cells(f"A{legend_row}:E{legend_row}")
leg_title = ws[f"A{legend_row}"]
leg_title.value = "Legend — Status Colors"
leg_title.font = Font(name="Calibri", bold=True, size=10, color=HEADER_FG)
leg_title.fill = PatternFill("solid", fgColor=HEADER_BG)
leg_title.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[legend_row].height = 16

legend_items = [
    ("IMPLEMENTED / QA PASS",              "E8F5E9", "1B5E20"),
    ("BACKEND-BLOCKED",                    "FFF3E0", "E65100"),
    ("GATE 3 COMPLETE — Awaiting Gate 4",  "E3F2FD", "0D47A1"),
    ("SUBSUMED (Implemented via CR-139)",  "F3E5F5", "4A148C"),
]
for i, (label, bg, fg) in enumerate(legend_items):
    lr = legend_row + 1 + i
    cell = ws.cell(row=lr, column=1, value=label)
    cell.font = Font(name="Calibri", size=9, bold=True, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    ws.row_dimensions[lr].height = 14

# ── Freeze panes ──────────────────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── Auto-filter on header row ─────────────────────────────────────────────────
ws.auto_filter.ref = f"A{header_row}:E{header_row + len(bugs)}"

output_path = "/app/memory/BUG_REPORT_AUG_07_TO_14_2026.xlsx"
wb.save(output_path)
print(f"Excel saved → {output_path}")
print(f"Total bugs: {len(bugs)}")
